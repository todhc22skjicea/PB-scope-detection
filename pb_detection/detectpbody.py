import argparse
import time
from pathlib import Path
import openpyxl
import cv2
import torch
import torch.backends.cudnn as cudnn
from numpy import random
import numpy as np
from brightness import calculate_brightness
from models.experimental import attempt_load
from utils.datasets import LoadStreams, LoadImages
from utils.general import check_img_size, check_requirements, check_imshow, non_max_suppression, apply_classifier, \
    scale_coords, xyxy2xywh, strip_optimizer, set_logging, increment_path
from utils.plots import plot_one_box
from utils.torch_utils import select_device, load_classifier, time_synchronized, TracedModel
import re  

def detect(save_img=False):
    weights, save_txt, imgsz=  opt.weights, opt.save_txt, opt.img_size
    save_img = True  # save inference images
    folderfile_number = 0
    # Directories
    save_dir = Path(increment_path(Path(opt.project) / opt.name, exist_ok=opt.exist_ok))  # increment run
    (save_dir / 'labels' if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir

    # Initialize
    set_logging()
    device = select_device(opt.device)
    half = device.type != 'cpu'  # half precision only supported on CUDA

    # Load model
    model = attempt_load(weights, map_location=device)  # load FP32 model
    stride = int(model.stride.max())  # model stride
    imgsz = check_img_size(imgsz, s=stride)  # check img_size

    #if trace:
    #    model = TracedModel(model, device, opt.img_size)

    if half:
        model.half()  # to FP16

    # Second-stage classifier
    #classify = False
    #if classify:
    #    modelc = load_classifier(name='resnet101', n=2)  # initialize
    #    modelc.load_state_dict(torch.load('weights/resnet101.pt', map_location=device)['model']).to(device).eval()

    # Set Dataloader
    vid_path, vid_writer = None, None
    print("image path",source)
    dataset = LoadImages(source, img_size=imgsz, stride=stride)

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names
    colors = [0,255,255]

    # Run inference
    if device.type != 'cpu':
        model(torch.zeros(1, 3, imgsz, imgsz).to(device).type_as(next(model.parameters())))  # run once
    old_img_w = old_img_h = imgsz
    old_img_b = 1

    t0 = time.time()
    for path, img, im0s, vid_cap in dataset:
        #print(img.dtype)
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)

        # Warmup
        if device.type != 'cpu' and (old_img_b != img.shape[0] or old_img_h != img.shape[2] or old_img_w != img.shape[3]):
            old_img_b = img.shape[0]
            old_img_h = img.shape[2]
            old_img_w = img.shape[3]
            for i in range(3):
                model(img, augment=opt.augment)[0]

        # Inference
        t1 = time_synchronized()
        with torch.no_grad():   # Calculating gradients would cause a GPU memory leak
            pred = model(img, augment=opt.augment)[0]
        t2 = time_synchronized()

        # Apply NMS
        pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, classes=opt.classes, agnostic=opt.agnostic_nms)
        t3 = time_synchronized()

        # Apply Classifier
        #if classify:
        #    pred = apply_classifier(pred, modelc, img, im0s)

        # Process detections
        for i, det in enumerate(pred):  # detections per image
            
            p, s, im0, frame = path, '', im0s, getattr(dataset, 'frame', 0)
            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # img.jpg
            #txt_path = str(save_dir)+'/labels'+'/detect.txt'  # img.txt
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            parts = p.name.rstrip('.png').split('_')      

            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(img.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum().item() # detections per class
                    #s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string
                #print(n)

                
                folderfile_number = folderfile_number+1
                file_number = file_index + folderfile_number
                image_path = str(p).replace('quantify','dataraw')
                npy_path = image_path.replace('pbody','mito').replace('C3.tif','C1_seg.npy').replace('T0008','T0001').replace('T0007','T0001').replace('T0006','T0001').replace('T0005','T0001').replace('T0004','T0001').replace('T0003','T0001').replace('T0002','T0001').replace('T8','T1').replace('T7','T1').replace('T6','T1').replace('T5','T1').replace('T4','T1').replace('T3','T1').replace('T2','T1')
                [aver_br, aver_bg, cell_area, cell_number] = calculate_brightness(image_path,npy_path)
                match = re.search('W(\d+)', image_path)
                group = int(match.group(1))
                drugname_index = 'A'+str(file_number)
                ws[drugname_index] = folders[np.where(group == index_e)[0][0]]
                time_index = 'B'+str(file_number)
                ws[time_index] = 'T'+str(t_index)
                filename_index = 'C'+str(file_number)  
                ws[filename_index] = str(p).replace(source+'/','')
                celln_index = 'D'+str(file_number)  
                ws[celln_index] = cell_number             
                pbodyn_index = 'E'+str(file_number)  
                ws[pbodyn_index] = n
                aver_pbody_index = 'F'+str(file_number) 
                if cell_number == 0:
                  ws[aver_pbody_index] = 'error'
                else:
                  ws[aver_pbody_index] = n/cell_number
                aver_bri_index = 'G'+str(file_number)  
                ws[aver_bri_index] = aver_br
                bg_bri_index = 'H'+str(file_number)  
                ws[bg_bri_index] = aver_bg
                cell_area_index = 'I'+str(file_number)
                ws[cell_area_index] = cell_area

                '''
                # Write results
                for *xyxy, conf, cls in reversed(det):
                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                        line = (cls, *xywh, conf) if opt.save_conf else (cls, *xywh)  # label format
                        with open(txt_path, 'a') as f:
                            f.write(('%g ' * len(line)).rstrip() % line + '\n')

                    if save_img:  # Add bbox to image
                        #label = f'{names[int(cls)]} {conf:.2f}'
                        plot_one_box(xyxy, im0, color=colors, line_thickness=1)
                '''
            # Print time (inference + NMS)
            print(f'{s}Done. ({(1E3 * (t2 - t1)):.1f}ms) Inference, ({(1E3 * (t3 - t2)):.1f}ms) NMS')

    return folderfile_number
    '''
    if save_txt or save_img:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ''
        print(f"Results saved to {save_dir}{s}")

    print(f'Done. ({time.time() - t0:.3f}s)')
    '''

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--weights', nargs='+', type=str, default='/home/detection/runs/train/pbody/weights/best.pt', help='model.pt path(s)')
    parser.add_argument('--source', type=str, default='inference/images', help='source')  # file/folder
    parser.add_argument('--img-size', type=int, default=4032, help='inference size (pixels)')
    #parser.add_argument('--source', type=str, default="", help='image path')
    parser.add_argument('--conf-thres', type=float, default=0.25, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.3, help='IOU threshold for NMS')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--view-img', action='store_true', help='display results')
    parser.add_argument('--save-txt', action='store_true', default=True, help='save results to *.txt')
    parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 0 2 3')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--update', action='store_true', help='update all models')
    parser.add_argument('--project', default='runs/detect', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    opt = parser.parse_args()
    print(opt)
    #check_requirements(exclude=('pycocotools', 'thop'))



    with torch.no_grad():

      for j in range(1,6):
        excel_name = "drug_screening_G"+str(j)+".xlsx"
        workbook = openpyxl.Workbook()
        workbook.save(excel_name) 
        wb = openpyxl.load_workbook(filename=excel_name)
        ws = wb.active
        file_index = 1
        ws['A1'] = 'Drugname'
        ws['B1'] = 'Time'
        ws['C1'] = 'Filename'
        ws['D1'] = 'Cell_number'
        ws['E1'] = 'Pbody_number'
        ws['F1'] = 'aver_pbody_number'
        ws['G1'] = 'aver_brightness'
        ws['H1'] = 'background_brightness'
        ws['I1'] = 'cell_area'
        #for t_index in [1,2,3,4,5,6,7,8,12]:    
        for t_index in range(1,9):
          source = "/media/tsuboilab/quantify/T"+str(t_index)+"/G"+str(j)+"/pbody"
          if j == 1:
            folders = ["T1648", "T1297", "T0437", "T0374", "T0429", "T0373", "T0461", "T0433", "T0364", "T0038","T0030", "T0129", "T0147", "T0163", "T0256", "T0304", "T0314", "T0327", "T0342", "T0492","T0498", "T0520", "T0263", "T0374L", "T2215", "T0973", "T1020", "T1056", "T1090", "T1144","T1146", "T1158", "T1159", "T0239", "T1291", "T2205", "T1181", "T1188", "T1210", "T1222","T0167", "T2597", "T0610", "T0646", "T2565", "T0711", "T0704", "T0692", "T0679", "T0773","T0678", "T0449", "T2508", "T0445", "T1636", "T2586","MG132","TG","DMSO","NA"]  #Group1
          if j == 2:
            folders = ["T0772", "T0740", "T0875", "T0891", "T2995", "T2827", "T2858", "T0800", "T0801", "T0809", "T0858", "T0860", "T2546", "T0392", "T0928", "T0933", "T1410", "T1418", "T1431", "T1439", "T1448", "T1452", "T1454", "T1477", "T1524", "T1537", "T1546", "T1558", "T1563", "T1661", "T1737", "T2175", "T2144", "T2399", "T1630", "T1621", "T1639", "T1642", "T1659", "T1660", "T1684", "T2381", "T2364", "T2382", "T2372", "T2369", "T2145", "T2115", "T2587", "T2532", "T1835", "T0335", "T8222", "T1085L", "T0152", "T2303", "MG132", "TG", "DMSO","NA"]
          if j == 3:
            folders = ["T2534", "T0097L", "T2148", "T2490", "T2483", "T3060", "T3059", "T1995", "T2325", "T2328", "T2920", "T2984", "T2851", "T6218", "T6230", "T3091", "T6227", "T2946", "T1656", "T1266", "T0080", "T0078", "T1038", "T1670", "T1829", "T1743", "T1777", "T1797", "T1912", "T1784", "T1785", "T1791", "T1792", "T1506", "T2220", "T0093L", "T2677", "T2456", "T2509", "T2539", "T2500", "T2066", "T2125", "T2485", "T2397", "T3067", "T2609", "T2656", "T1963", "T1921", "T6199", "T8387", "T14998", "T6460", "T1260", "T8151", "MG132", "TG", "DMSO","NA"] 
          if j == 4:
            folders = ["T1929", "T1936", "T1975", "T1894", "T1903", "T1988", "T3061", "T3211", "T6321", "T3269", "T3616", "T3402", "T3380", "T3626", "T3625", "T6115", "T3623", "T3634", "T3678", "T6165", "T6019", "T3726", "T6867", "T6121", "T6120", "T6280", "T6302", "T6758", "T6345", "T2S0007", "T6169", "T2P2923", "T6723", "T6588", "T4168", "T6156", "T6020", "T4332", "T7503", "T6880", "T6101", "T4575", "T4409", "T6487", "T6674", "T4749", "T4976", "T5001", "T5030", "T5109", "T3603", "T1633", "T0745", "T2796", "T7175", "T3O2749", "MG132", "TG", "DMSO","NA"]
          if j == 5:
            folders = ["T5171", "T5177", "T5462", "T5882", "T7094", "T6930", "T5995", "T7584", "T7394", "T7486", "T7861", "T8132", "T5857", "T7604", "T8402", "T6914", "T8399", "T6475", "TQ0277", "T1791L", "T8482", "T12401", "TQ0210", "T8474", "TQ0319", "T8541", "T15732", "T8651", "T8654", "T2147", "T12317", "T8684", "T8825", "T22235", "TQ0064", "T0979", "T10358", "T19965", "T12311", "T1756L", "T22324", "T10585", "T20029", "T12594", "T15675", "T0247", "T1035", "T0878", "T2854", "T5016", "T4883", "T0194", "T2211", "T0033", "T13202","NA", "MG132", "TG", "DMSO","NA"] 
          index_e = np.array([14,15,16,17,18,19,20,21,22,23,26,27,28,29,30,31,32,33,34,35,38,39,40,41,42,43,44,45,46,47,50,51,52,53,54,55,56,57,58,59,62,63,64,65,66,67,68,69,70,71,74,75,76,77,78,79,80,81,82,83])   
          folderfile_number = detect()
 
          file_index = file_index+folderfile_number
        wb.save(excel_name)
          
          
